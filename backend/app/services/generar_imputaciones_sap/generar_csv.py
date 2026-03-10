# PATH: backend/app/services/generar_imputaciones_sap/generar_csv.py

import os
import csv
import pandas as pd
from datetime import datetime
from sqlalchemy.orm import Session
from app.models.models import TablaCentral
from zipfile import ZipFile
from openpyxl.utils import get_column_letter

def fetch_data(db: Session) -> pd.DataFrame:
    query = db.query(
        TablaCentral.Employee_Number,
        TablaCentral.Date,
        TablaCentral.HourType,
        TablaCentral.ProductionOrder,
        TablaCentral.Operation,
        TablaCentral.OperationActivity,
        TablaCentral.Hours
    ).filter(TablaCentral.Cargado_SAP == False)

    result = db.execute(query.statement)
    df = pd.DataFrame(result.fetchall(), columns=result.keys())
    return df

def map_hourtype(op_act, original_hourtype):
    if isinstance(op_act, str):
        if op_act.endswith("XX"):
            return 3
        if op_act.endswith("GG"):
            return 4
        if len(op_act) >= 2 and op_act[-2] == "C":
            return 5
    return original_hourtype

def generate_zip_with_csv_and_xlsx(db: Session) -> str:
    """
    Genera un ZIP con dos ficheros (CSV + XLSX) a partir de las filas de TablaCentral
    donde Cargado_SAP == False.

    •  El CSV usa ‘;’ como separador y codificación cp1252 (idéntico a antes).
    •  Se **elimina** la comilla simple que se añadía delante de
        ProductionOrder, Operation y OperationActivity.
    """
    data = fetch_data(db)
    if data.empty:
        return None

    # 1) Asegurar que Date sea datetime64 para formateo correcto en CSV y XLSX
    data["Date"] = pd.to_datetime(data["Date"])

    # 2) Columnas extra vacías
    for col in [
        "Project",
        "Wbs",
        "Cost Center",
        "Activity Type",
        "Status",
        "Serial Number",
    ]:
        data[col] = ""

    # 3) Ajustar HourType según OperationActivity
    data["HourType"] = [
        map_hourtype(a, h) for a, h in zip(data["OperationActivity"], data["HourType"])
    ]

    # 4) Columnas finales y reordenación
    cols_final = [
        "Employee_Number",
        "Date",
        "HourType",
        "Project",
        "Wbs",
        "Cost Center",
        "Activity Type",
        "ProductionOrder",
        "Operation",
        "OperationActivity",
        "Hours",
        "Status",
        "Serial Number",
    ]
    data_final = data[cols_final]

    # 5) Rutas temporales
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"mass_upload_{timestamp}"
    tmp_dir = os.path.join(os.getcwd(), "tmp_csv_sap")
    os.makedirs(tmp_dir, exist_ok=True)

    csv_path = os.path.join(tmp_dir, base_name + ".csv")
    xlsx_path = os.path.join(tmp_dir, base_name + ".xlsx")
    zip_path = os.path.join(tmp_dir, base_name + ".zip")

    # 6) Guardar CSV con fechas dd/mm/YYYY
    data_final.to_csv(
        csv_path,
        sep=";",
        encoding="cp1252",
        index=False,
        lineterminator="\r\n",
        quoting=csv.QUOTE_MINIMAL,
        date_format="%d/%m/%Y",
    )

    # 7) Guardar XLSX con fechas como celdas de fecha reales (formato DD/MM/YYYY)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl",
                        date_format="DD/MM/YYYY",
                        datetime_format="DD/MM/YYYY") as writer:
        data_final.to_excel(writer, index=False)
        ws = writer.sheets["Sheet1"]
        # Buscar la columna "Date" y forzar el number_format en todas sus celdas
        date_col_idx = cols_final.index("Date") + 1  # 1-based
        date_col_letter = get_column_letter(date_col_idx)
        for cell in ws[date_col_letter][1:]:  # skip header
            cell.number_format = "DD/MM/YYYY"

    # 8) Empaquetar ZIP
    with ZipFile(zip_path, "w") as z:
        z.write(csv_path, arcname=os.path.basename(csv_path))
        z.write(xlsx_path, arcname=os.path.basename(xlsx_path))

    return zip_path
