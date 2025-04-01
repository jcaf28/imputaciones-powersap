# PATH: backend/app/services/utils/escaneos_utils.py

import os
import pandas as pd
import glob
import subprocess
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font


def auto_ajustar_ancho_columnas(ws):
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

def aplicar_estilos_excel(ruta_archivo):
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    auto_ajustar_ancho_columnas(ws)

    # Estilo de la primera fila y fijarla
    primera_fila = ws[1]
    for celda in primera_fila:
        celda.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    ws.freeze_panes = "A2"

    # Aplicar estilo de relleno alternativo más sutil
    color_fill_1 = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
    color_fill_2 = PatternFill(fill_type=None)
    fill_actual = color_fill_1
    orden_anterior = None

    for fila in ws.iter_rows(min_row=2):
        orden_actual = fila[0].value
        if orden_actual != orden_anterior:
            fill_actual = color_fill_1 if fill_actual == color_fill_2 else color_fill_2
        for celda in fila:
            if celda.value is not None:
                celda.fill = fill_actual
        orden_anterior = orden_actual

    # Encontrar número de columna para 'Ubicación PDF'
    columna_ubicacion_pdf = encontrar_numero_columna(ws, 'UbicacionPDF')

    # Convertir UbicacionPDF en hipervínculos
    if columna_ubicacion_pdf:
        for row in ws.iter_rows(min_row=2, min_col=columna_ubicacion_pdf, max_col=columna_ubicacion_pdf):
          for cell in row:
              cell.hyperlink = cell.value
              cell.font = Font(underline="single", color="0563C1")

    # Agregar formato condicional para 'OA SFI Status'
    columna_oa_sfi_status = encontrar_numero_columna(ws, 'OA SFI Status')
    if columna_oa_sfi_status:
        for row in ws.iter_rows(min_row=2, min_col=columna_oa_sfi_status, max_col=columna_oa_sfi_status):
            for cell in row:
                if cell.value == 'In Queue':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif cell.value == 'In Process':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif cell.value in ['Completed with pending work', 'Completed with missing parts']:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif cell.value == 'Completed':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    
    wb.save(ruta_archivo)

def obtener_nombres_pdf(ruta_carpeta):
    pdfs = []
    for raiz, dirs, archivos in os.walk(ruta_carpeta):
        for archivo in archivos:
            if archivo.endswith('.pdf'):
                ruta_completa = os.path.join(raiz, archivo)
                pdfs.append((archivo, ruta_completa))
    return pdfs

def leer_xlsx_mas_reciente(ruta_carpeta):
    lista_archivos = glob.glob(ruta_carpeta + '/*.xlsx')
    ultimo_archivo = max(lista_archivos, key=os.path.getctime)
    return pd.read_excel(ultimo_archivo)

def abrir_explorador(ruta):
    ruta = os.path.normpath(ruta)
    print(f"Abriendo explorador de archivos en: {ruta}")
    subprocess.Popen(f'explorer /select,{ruta}')

def procesar_dataframe_encontrados(df):
    # Asegurar que 'Order' y 'Operation Activity' sean las primeras columnas y quitar los paréntesis
    df['Order'] = df['Order'].str.replace(r'[()]', '', regex=True)
    df['Operation Activity'] = df['Operation Activity'].str.split('-').str[0]  # Asegurarse de que solo se use la primera parte antes del guión
    columnas = ['Order', 'Operation Activity'] + [col for col in df.columns if col not in ['Order', 'Operation Activity']]
    df = df[columnas]
    df.sort_values(by=['Order', 'Operation Activity'], inplace=True)
    return df

def encontrar_numero_columna(ws, nombre_columna):
    for i, columna in enumerate(ws.iter_cols(min_row=1, max_row=1)):
        if columna[0].value == nombre_columna:
            return i + 1  # Las columnas en openpyxl comienzan en 1, no en 0
    return None

# Inicializa el nombre del archivo de log al comienzo del script
timestamp_log = datetime.now().strftime("%Y%m%d_%H%M%S")
nombre_archivo_log = f"log_cierres_{timestamp_log}.txt"

def escribir_log(mensaje, ruta_xlsx):
    directorio_logs = os.path.join(ruta_xlsx, 'logs')
    os.makedirs(directorio_logs, exist_ok=True)  # Crea el directorio si no existe
    ruta_archivo_log = os.path.join(directorio_logs, nombre_archivo_log)

    # Escribir el mensaje en el archivo de log
    with open(ruta_archivo_log, 'a') as archivo_log:
        archivo_log.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {mensaje}\n")

    # Además, imprime el mensaje en la consola
    print(mensaje)

