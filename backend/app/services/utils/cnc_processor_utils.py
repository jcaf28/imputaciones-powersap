# PATH: backend/app/services/utils/cnc_processor_utils.py

import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from openpyxl.utils import get_column_letter
from typing import Dict

# ========== FUNCIONES AUXILIARES ==========

def encontrar_archivo_mas_reciente():
    """
    Busca el archivo de avisos de calidad más reciente en la carpeta 'backend/tmp'.
    Se asegura de calcular la ruta absoluta correctamente sin importar si está dentro o fuera de Docker.
    """
    # Obtener la ruta absoluta de la carpeta 'backend/tmp'
    base_dir = Path(__file__).resolve().parent.parent.parent.parent  # Retrocede desde /backend/app/utils/
    
    # Si el script se ejecuta dentro de Docker, `base_dir` podría ser `/app`, por lo que forzamos `backend/tmp`
    if base_dir.name == "app":
        backend_dir = base_dir.parent  # Esto debería ser `/backend`
    else:
        backend_dir = base_dir

    directorio = backend_dir / "tmp"

    # Buscar archivos en ese directorio
    lista_archivos = list(directorio.glob("*.xlsx"))
    if not lista_archivos:
        raise FileNotFoundError(f"No se encontraron archivos de avisos de calidad en {directorio}")

    # Obtener el archivo más reciente
    archivo_mas_reciente = max(lista_archivos, key=os.path.getmtime)
    return str(archivo_mas_reciente)  # Devolver la ruta absoluta como string

def extraer_avisos_de_calidad(archivo_avisos: str) -> Dict[str, Dict[str, list]]:
    df_avisos = pd.read_excel(archivo_avisos, dtype=str)
    avisos_calidad = {'C1': {}, 'C2': {}, 'C3': {}}

    for _, row in df_avisos.iterrows():
        qn_type = row.get("QN Type") or row.get("Clase de aviso")
        qn_number = row.get("QN Number") or row.get("Número de notificación")
        short_text = str(row.get("Short text of QN Header") or row.get("Texto breve"))
        modelo = short_text.split(' ')[-1] if short_text else None

        if qn_type and modelo and qn_number:
            if qn_type == 'Y1':
                avisos_calidad['C1'].setdefault(modelo, []).append(qn_number)
            elif qn_type == 'Y5':
                avisos_calidad['C2'].setdefault(modelo, []).append(qn_number)
            elif qn_type == 'Y8':
                avisos_calidad['C3'].setdefault(modelo, []).append(qn_number)

    return avisos_calidad

def aplicar_formato(ws):
    color_fill_1 = PatternFill(start_color='EFEFEF', end_color='EFEFEF', fill_type='solid')  # Gris claro
    color_fill_2 = PatternFill(start_color='DFDFDF', end_color='DFDFDF', fill_type='solid')  # Gris oscuro

    fill_actual = color_fill_1
    modelo_anterior = None
    columna_modelo_index = 1  # 'MODELO' es la 1ª columna

    # Ajustar el ancho de las columnas y aplicar estilos a cabecera
    for col_num, column in enumerate(ws.columns, 1):
        max_length = len(str(ws.cell(row=1, column=col_num).value))
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    # Formato zebra por modelo
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=False):
        modelo_actual = row[0].value
        if modelo_actual != modelo_anterior:
            fill_actual = color_fill_2 if fill_actual == color_fill_1 else color_fill_1
            modelo_anterior = modelo_actual
        for cell in row:
            cell.fill = fill_actual

    # Cabecera en negrita
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # Fijar la primera fila
    ws.freeze_panes = 'A2'

def aplicar_formato_condicional(ws):
    for row in ws.iter_rows(min_row=2, values_only=False):
        modelo = row[0].value  # columna 'MODELO'
        for cell in row:
            # columnas que empiezan por 'CANT.':
            encabezado = ws.cell(row=1, column=cell.column).value or ""
            if encabezado.startswith('CANT.'):
                cell.alignment = Alignment(horizontal="center")

                if encabezado == 'CANT. C0s':
                    # Formato especial para C0 según sea 'EA' o 'EB'
                    objetivo = 7 if 'EA' in modelo else 15 if 'EB' in modelo else None
                    if objetivo:
                        if cell.value == objetivo:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value > objetivo:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif 0 < cell.value < objetivo:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    # C1, C2, C3
                    if cell.value == 1:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif cell.value > 1:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def ordenar_registros(ws):
    registros = []
    # Copiamos todos los valores (excepto cabecera) en memoria
    for row in ws.iter_rows(min_row=2, values_only=True):
        registros.append(row)

    cabecera = [cell.value for cell in ws[1]]
    modelo_idx = cabecera.index('MODELO')
    unidad_idx = cabecera.index('UNIDAD')

    def clave_ordenacion(registro):
        modelo = registro[modelo_idx]
        try:
            unidad = int(registro[unidad_idx])
        except ValueError:
            unidad = float('inf')
        return (modelo, unidad)

    registros_ordenados = sorted(registros, key=clave_ordenacion)

    # Borramos filas y reescribimos en orden
    ws.delete_rows(2, ws.max_row - 1)
    for idx, registro in enumerate(registros_ordenados, start=2):
        for jdx, valor in enumerate(registro):
            ws.cell(row=idx, column=jdx+1, value=valor)