# PATH: backend/app/services/feedback/feedback_processor.py

import pandas as pd
import tempfile, os
import openpyxl
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.db.session import database_session
from app.models.models import Imputaciones, TablaCentral
from app.core.sse_manager import sse_manager
from app.services.feedback._utils import change_dtypes, intercambiar_tareas, _filtrar_fechas_parseables, is_null

# ==== Procesamiento Completo ====

def procesar_feedback_completo(
    df: pd.DataFrame,
    process_id: str,
    original_filename: str
):
    sse_manager.send_message(process_id, "🔍 Filtrando fechas parseables…")
    df = _filtrar_fechas_parseables(df)

    sse_manager.send_message(process_id, "⚙️ Ajustando tipos de datos…")
    df = change_dtypes(df)

    sse_manager.send_message(process_id, "🔄 Intercambiando tareas…")
    df = intercambiar_tareas(df)

    # columnas resultado
    for col in (
        "Estado",
        "Imputacion_ID",
        "SAP_Order",
        "SAP_OperationActivity",
        "SAP_EffectivityFull",
    ):
        df[col] = None

    sse_manager.send_message(process_id, "📡 Consultando la base de datos…")

    with database_session as db:
        for idx, row in df.iterrows():
            estado, imp_id, sap_ord, sap_opact, sap_eff = _obtener_estado_imputacion(
                db, df, row
            )
            df.loc[idx, ["Estado",
                         "Imputacion_ID",
                         "SAP_Order",
                         "SAP_OperationActivity",
                         "SAP_EffectivityFull"]] = (
                estado,
                imp_id,
                sap_ord,
                sap_opact,
                sap_eff,
            )
            if idx % 10 == 0:
                sse_manager.send_message(
                    process_id, f"🔢 Procesadas {idx + 1}/{len(df)} filas…"
                )

    sse_manager.send_message(process_id, "🎨 Aplicando colores…")

    # -------- nombre de salida: feedback_<nombre-entrada>.xlsx ----------
    base = os.path.splitext(os.path.basename(original_filename))[0]
    temp_file = tempfile.NamedTemporaryFile(
        delete=False, prefix=f"feedback_{base}_", suffix=".xlsx"
    )
    df.to_excel(temp_file.name, index=False)
    _colorear_celdas_según_estado(temp_file.name)

    sse_manager.send_message(process_id, "✅ Archivo generado correctamente.")
    return temp_file.name, f"feedback_{base}.xlsx"

# ==== Funciones Auxiliares ====

def _obtener_estado_imputacion(db: Session, df: pd.DataFrame, row: pd.Series):
    try:
        fecha_imp = pd.to_datetime(row['FechaImp']).date() if not is_null(row['FechaImp']) else None
        cod_empleado = str(int(row['CodEmpleado'])) if not is_null(row['CodEmpleado']) else None
        timpu = str(int(row['Timpu'])) if not is_null(row['Timpu']) else None
        proyecto = str(row['Proyecto']) if not is_null(row['Proyecto']) else None
        tipo_coche = str(row['TipoCoche']) if not is_null(row['TipoCoche']) else None
        num_coche = str(int(float(row['NumCoche']))) if not is_null(row['NumCoche']) else None
        centro_trabajo = str(int(row['CentroTrabajo'])) if not is_null(row['CentroTrabajo']) else None
        tarea = str(row['Tarea']) if not is_null(row['Tarea']) else None
        tarea_asoc = str(row['TareaAsoc']) if not is_null(row['TareaAsoc']) else None
        horas = round(float(row['Horas']), 2) if not is_null(row['Horas']) else None
    except Exception as e:
        return "0 - No admitida", None, None, None, None

    ids_ya_asignados = set(df['Imputacion_ID'].dropna().unique())

    query = db.query(Imputaciones).filter(
        Imputaciones.FechaImp == fecha_imp,
        Imputaciones.CodEmpleado == cod_empleado,
        Imputaciones.Timpu == timpu,
        Imputaciones.Proyecto == proyecto,
        Imputaciones.TipoCoche == tipo_coche,
        Imputaciones.NumCoche == num_coche,
        Imputaciones.CentroTrabajo == centro_trabajo,
        Imputaciones.Tarea == tarea,
        Imputaciones.TareaAsoc == tarea_asoc,
        Imputaciones.Horas == horas
    )

    imputacion = next((imp for imp in query.all() if imp.ID not in ids_ya_asignados), None)

    if not imputacion:
        return "0 - No admitida", None, None, None, None

    tabla_central = db.query(TablaCentral).filter(TablaCentral.imputacion_id == imputacion.ID).first()

    if not tabla_central:
        return "0 - No admitida", imputacion.ID, None, None, None

    sap_order = tabla_central.sap_order
    sap_order_val = sap_order.Order if sap_order else None
    sap_opact_val = sap_order.OperationActivity if sap_order else None
    sap_eff_val = sap_order.EffectivityFull if sap_order else None

    if tabla_central.Cargado_SAP and tabla_central.cargadoEnTareaReal:
        estado = "1- Cargado correctamente en SAP"
    elif tabla_central.Cargado_SAP and not tabla_central.cargadoEnTareaReal:
        estado = "2- Cargado en SAP a una tarea alternativa"
    elif not tabla_central.Cargado_SAP and tabla_central.cargadoEnTareaReal:
        estado = "3a- tarea encontrada pero imputación no cargada en sap"
    else:
        estado = "3b- tarea alternativa encontrada pero imputación no cargada en sap"

    return estado, imputacion.ID, sap_order_val, sap_opact_val, sap_eff_val

def _colorear_celdas_según_estado(ruta_archivo):
    wb = openpyxl.load_workbook(ruta_archivo)
    ws = wb.active

    col_estado = None
    for col in ws.iter_cols(min_row=1, max_col=ws.max_column, max_row=1):
        for cell in col:
            if cell.value == "Estado":
                col_estado = cell.column
                break
        if col_estado:
            break

    colores = {
        "0 - No admitida": "C00000",  # rojo
        "1- Cargado correctamente en SAP": "00B050",  # verde oscuro
        "2- Cargado en SAP a una tarea alternativa": "92D050",  # verde claro
        "3a- tarea encontrada pero imputación no cargada en sap": "FFC000",  # ámbar
        "3b- tarea alternativa encontrada pero imputación no cargada en sap": "FFC000"  # ámbar
    }

    if col_estado:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            estado = ws.cell(row=row[0].row, column=col_estado).value
            color = colores.get(estado, "000000")
            for cell in row:
                cell.font = Font(color=color)

    wb.save(ruta_archivo)
