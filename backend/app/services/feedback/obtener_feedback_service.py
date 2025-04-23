# PATH: backend/app/services/feedback/obtener_feedback_service.py

import pandas as pd, openpyxl
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.db.session import database_session
from app.models.models import Imputaciones, TablaCentral
from app.services.feedback._utils import change_dtypes, intercambiar_tareas
from app.core.sse_manager import sse_manager


# --------------------------------------------------------------------------- #
# 1) PREPARAR DF                                                               #
# --------------------------------------------------------------------------- #
def preparar_dataframe_feedback(df: pd.DataFrame) -> pd.DataFrame:
    if "Fecha" in df.columns:
        df = df.rename(columns={"Fecha": "FechaImp"})

    df = filtrar_fechas_parseables(df)
    df.replace(["None", "nan"], None, inplace=True)

    df = change_dtypes(df)
    df = intercambiar_tareas(df)
    df.replace(["None", "nan"], None, inplace=True)

    for col in [
        "Estado",
        "Imputacion_ID",
        "SAP_Order",
        "SAP_OperationActivity",
        "SAP_EffectivityFull",
    ]:
        df[col] = None

    return df


# --------------------------------------------------------------------------- #
# 2) GENERAR EXCEL EN MEMORIA                                                  #
# --------------------------------------------------------------------------- #
def generar_xlsx_en_memoria(
    df: pd.DataFrame, original_name: str, process_id: str | None = None
) -> tuple[str, bytes]:
    asignar_estados(df, process_id)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    buffer.seek(0)

    # colorear
    wb = openpyxl.load_workbook(buffer)
    ws = wb.active
    col_estado = next((c.column for c in ws[1] if c.value == "Estado"), None)
    if col_estado:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            estado = ws.cell(row=row[0].row, column=col_estado).value
            color = obtener_color_estado(estado)
            for cell in row:
                cell.font = Font(color=color)

    out_buffer = BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)

    # nombre sugerido de descarga
    name, ext = original_name.rsplit(".", 1)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{name}_feedback_{ts}.xlsx"

    return filename, out_buffer.read()


# ================== auxiliares (idénticos a antes) ==========================
def filtrar_fechas_parseables(df):
    return df[df["FechaImp"].apply(lambda x: _es_fecha(x))]


def _es_fecha(x):
    try:
        pd.to_datetime(x)
        return True
    except Exception:
        return False


def asignar_estados(df: pd.DataFrame, process_id: str | None = None):
    with database_session as db:
        for i, row in df.iterrows():
            est, imp_id, o, opact, eff = obtener_estado_imputacion(db, df, row)
            df.at[i, "Estado"] = est
            df.at[i, "Imputacion_ID"] = imp_id
            df.at[i, "SAP_Order"] = o
            df.at[i, "SAP_OperationActivity"] = opact
            df.at[i, "SAP_EffectivityFull"] = eff

            if process_id and (i + 1) % 200 == 0:
                sse_manager.send_message(
                    process_id, f"🔎 {i+1}/{len(df)} filas procesadas…"
                )


def obtener_estado_imputacion(
    db: Session, df: pd.DataFrame, row: pd.Series
):  # igual que antes
    from datetime import datetime

    sap_order_val = sap_opact_val = sap_eff_val = None
    try:
        if isinstance(row["FechaImp"], datetime):
            fecha_imp = row["FechaImp"].date()
        elif pd.notna(row["FechaImp"]):
            fecha_imp = datetime.strptime(row["FechaImp"], "%d/%m/%Y").date()
        else:
            fecha_imp = None

        cod_emp = str(int(row["CodEmpleado"])) if pd.notna(row["CodEmpleado"]) else None
        timpu = str(int(row["Timpu"])) if pd.notna(row["Timpu"]) else None
        proj = str(row["Proyecto"]) if pd.notna(row["Proyecto"]) else None
        tipo_coche = str(row["TipoCoche"]) if pd.notna(row["TipoCoche"]) else None
        num_coche = (
            str(int(float(row["NumCoche"]))) if pd.notna(row["NumCoche"]) else None
        )
        centro = str(int(row["CentroTrabajo"])) if pd.notna(row["CentroTrabajo"]) else None
        tarea = str(row["Tarea"]) if pd.notna(row["Tarea"]) else None
        tarea_asoc = str(row["TareaAsoc"]) if pd.notna(row["TareaAsoc"]) else None
        horas = round(float(row["Horas"]), 2) if pd.notna(row["Horas"]) else None
    except ValueError:
        return "0 - No admitida", None, None, None, None

    ids_ya = set(df["Imputacion_ID"].dropna().unique())
    imp = (
        db.query(Imputaciones)
        .filter(
            Imputaciones.FechaImp == fecha_imp,
            Imputaciones.CodEmpleado == cod_emp,
            Imputaciones.Timpu == timpu,
            Imputaciones.Proyecto == proj,
            Imputaciones.TipoCoche == tipo_coche,
            Imputaciones.NumCoche == num_coche,
            Imputaciones.CentroTrabajo == centro,
            Imputaciones.Tarea == tarea,
            Imputaciones.TareaAsoc == tarea_asoc,
            Imputaciones.Horas == horas,
        )
        .all()
    )
    imputacion = next((x for x in imp if x.ID not in ids_ya), None)
    if not imputacion:
        return "0 - No admitida", None, None, None, None

    tabla = (
        db.query(TablaCentral)
        .filter(TablaCentral.imputacion_id == imputacion.ID)
        .first()
    )
    if not tabla:
        return "0 - No admitida", imputacion.ID, None, None, None

    if tabla.sap_order:
        sap_order_val = tabla.sap_order.Order
        sap_opact_val = tabla.sap_order.OperationActivity
        sap_eff_val = tabla.sap_order.EffectivityFull

    if tabla.Cargado_SAP and tabla.cargadoEnTareaReal:
        estado = "1- Cargado correctamente en SAP"
    elif tabla.Cargado_SAP and not tabla.cargadoEnTareaReal:
        estado = "2- Cargado en SAP a una tarea alternativa"
    elif not tabla.Cargado_SAP and tabla.cargadoEnTareaReal:
        estado = "3a- tarea encontrada pero imputación no cargada en sap"
    elif not tabla.Cargado_SAP and not tabla.cargadoEnTareaReal:
        estado = "3b- tarea alternativa encontrada pero imputación no cargada en sap"
    else:
        estado = "0 - No admitida"

    return estado, imputacion.ID, sap_order_val, sap_opact_val, sap_eff_val


def obtener_color_estado(estado: str):
    return {
        "0 - No admitida": "C00000",
        "1- Cargado correctamente en SAP": "00B050",
        "2- Cargado en SAP a una tarea alternativa": "92D050",
        "3a- tarea encontrada pero imputación no cargada en sap": "FFC000",
        "3b- tarea alternativa encontrada pero imputación no cargada en sap": "FFC000",
    }.get(estado, "000000")
