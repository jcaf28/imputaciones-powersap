# PATH: backend/app/services/file_processor.py

import time
from openpyxl import load_workbook

def process_file(original_filename: str, content: bytes) -> str:
  # Guardar el archivo original
  with open(original_filename, 'wb') as f:
    f.write(content)
  
  # Procesar el archivo de juguete
  wb = load_workbook(original_filename)
  ws = wb.active
  ws['A1'] = 'OK'
  
  timestamp = int(time.time())
  new_filename = f"{original_filename.split('.xlsx')[0]}_{timestamp}.xlsx"
  wb.save(new_filename)
  
  return new_filename
